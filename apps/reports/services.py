# apps/reports/services.py
from io import BytesIO
import csv
import json
import logging
from typing import Tuple, List, Dict, Any

from django.db.models import Sum, F
from django.utils import timezone

from apps.products.models import Product, ProductVariant
from apps.users.models import AuditLog

# billing models (pueden existir)
try:
    from apps.billing.models import Invoice, InvoiceItem, Reservation
except Exception:
    Invoice = None
    InvoiceItem = None
    Reservation = None

# pandas optional
try:
    import pandas as pd
except Exception:
    pd = None

logger = logging.getLogger(__name__)


class ReportService:
    """
    Registry-based report runner.
    Handlers registered with @ReportService.register("<report_type>").
    Each handler returns (rows: List[dict], columns: List[str]).
    """
    registry: Dict[str, Any] = {}

    @classmethod
    def register(cls, name: str):
        def _inner(fn):
            cls.registry[name] = fn
            return fn
        return _inner

    @classmethod
    def run(cls, report_definition: Any, params: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
        rtype = getattr(report_definition, "report_type", None) or report_definition
        handler = cls.registry.get(rtype)
        if handler is None:
            raise ValueError(f"No hay handler registrado para: {rtype!r}")
        return handler(report_definition, params or {})

    # ---- Export helpers ----
    @classmethod
    def to_xlsx_bytes(cls, rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.drawing.image import Image as XLImage
        import os

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Logo en Excel
        try:
            from django.conf import settings
            logo_path = os.path.join(settings.BASE_DIR, "static", "img", "Logo sin fondo azul.png")
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width, img.height = 150, 50
                ws.add_image(img, "A1")
                ws.append([])  # fila vacía para espaciar
                ws.append([])
        except Exception as e:
            logger.warning("No se pudo agregar logo en Excel: %s", e)

        # Encabezados
        header_fill = PatternFill("solid", fgColor="0d6efd")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(columns)
        for col_num, col_name in enumerate(columns, 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        for row in rows:
            ws.append([row.get(c, "") for c in columns])

        # Ajustar ancho de columnas
        for col_num, col_name in enumerate(columns, 1):
            ws.column_dimensions[chr(64 + col_num)].width = max(12, len(col_name) + 2)

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    @classmethod
    def to_csv_bytes(cls, rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
        bio = BytesIO()
        writer = csv.writer(bio)
        writer.writerow(columns)
        for r in rows:
            writer.writerow([r.get(c, "") for c in columns])
        return bio.getvalue()

    @classmethod
    def to_json_bytes(cls, rows: List[Dict[str, Any]]) -> bytes:
        return json.dumps(rows, default=str, ensure_ascii=False).encode("utf-8")

    @classmethod
    def to_pdf_bytes(cls, rows: List[Dict[str, Any]], columns: List[str], title: str = "Reporte") -> bytes:
        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, Image
            )
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
        except Exception as e:
            logger.error("Reportlab no disponible: %s", e)
            raise

        bio = BytesIO()
        doc = SimpleDocTemplate(bio, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        story = []

        # Encabezado con logo
        try:
            from django.conf import settings
            import os
            logo_path = os.path.join(settings.BASE_DIR, "static", "img", "Logo sin fondo azul.png")
            if os.path.exists(logo_path):
                img = Image(logo_path, width=120, height=40)
                story.append(img)
                story.append(Spacer(1, 12))
        except Exception as e:
            logger.warning("No se pudo cargar el logo: %s", e)

        # Título
        story.append(Paragraph(title, styles["Title"]))
        story.append(Spacer(1, 24))

        # Tabla
        data = [columns] + [[str(r.get(c, "")) for c in columns] for r in rows]
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0d6efd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)

        # Pie de página
        story.append(Spacer(1, 24))
        story.append(Paragraph(f"Generado en {timezone.now():%Y-%m-%d %H:%M}", styles["Normal"]))

        doc.build(story)
        return bio.getvalue()


# ---------------- Handlers ----------------

@ReportService.register("inventory")
def inventory_report(definition: Any, params: Dict[str, Any]):
    """
    Inventory: returns product stock + reserved + available + category + price.
    Filters supported: absolute_category (id), min_stock (int).
    """
    qs = Product.objects.all().select_related("absolute_category").prefetch_related("variants")

    if params.get("absolute_category"):
        try:
            qs = qs.filter(absolute_category_id=int(params["absolute_category"]))
        except Exception:
            pass

    min_stock = params.get("min_stock")
    if min_stock is not None:
        try:
            m = int(min_stock)
            qs = [p for p in qs if (p.stock or 0) >= m]
        except Exception:
            pass

    rows = []
    for p in qs:
        rows.append({
            "id": p.id,
            "sku": p.sku,
            "name": p.name,
            "category": getattr(p.absolute_category, "nombre", None) or getattr(p.absolute_category, "name", None),
            "stock": p.stock,
            "reserved_stock": p.reserved_stock,
            "available_stock": p.available_stock,
            "min_stock": p.min_stock,
            "price": float(p.price) if p.price is not None else None,
        })

    columns = ["id", "sku", "name", "category", "stock", "reserved_stock", "available_stock", "min_stock", "price"]
    return rows, columns


@ReportService.register("sales")
def sales_report(definition: Any, params: Dict[str, Any]):
    """
    Sales: listado de facturas.
    Parámetros: date_from, date_to (ISO), payment_method (optional).
    """
    if Invoice is None:
        raise RuntimeError("El modelo Invoice no está disponible en este proyecto.")

    qs = Invoice.objects.all().select_related("reservation")
    date_from = params.get("date_from")
    date_to = params.get("date_to")
    if date_from and date_to:
        try:
            qs = qs.filter(created_at__range=[date_from, date_to])
        except Exception:
            logger.warning("No se pudo aplicar date range %s - %s", date_from, date_to)

    if params.get("payment_method"):
        qs = qs.filter(payment_method=params["payment_method"])

    rows = []
    for inv in qs.order_by("-created_at"):
        items_count = getattr(inv, "items", None).count() if hasattr(inv, "items") else None
        reservation_id = inv.reservation.pk if getattr(inv, "reservation", None) else None
        rows.append({
            "id": inv.pk,
            "code": inv.code,
            "client": f"{inv.client_first_name or ''} {inv.client_last_name or ''}".strip(),
            "total": float(inv.total or 0),
            "subtotal": float(inv.subtotal or 0),
            "discount_amount": float(inv.discount_amount or 0),
            "amount_paid": float(inv.amount_paid or 0),
            "payment_method": inv.get_payment_method_display() if hasattr(inv, "get_payment_method_display") else inv.payment_method,
            "payment_provider": inv.payment_provider,
            "paid": bool(inv.paid),
            "status": inv.status,
            "items_count": items_count,
            "reservation_id": reservation_id,
            "created_at": inv.created_at.isoformat() if getattr(inv, "created_at", None) else None,
        })
    columns = ["id", "code", "client", "subtotal", "discount_amount", "total", "amount_paid", "payment_method", "payment_provider", "paid", "status", "items_count", "reservation_id", "created_at"]
    return rows, columns


@ReportService.register("top_products")
def top_products_report(definition: Any, params: Dict[str, Any]):
    """
    Top products report.
    Parámetros:
      - date_from, date_to (ISO)
      - limit (int, default 20)
      - mode ("top"=más vendidos, "bottom"=menos vendidos)
    """
    if InvoiceItem is None:
        raise RuntimeError("InvoiceItem no existe en este proyecto.")

    qs = InvoiceItem.objects.select_related("product", "invoice")

    # Filtrar por fechas
    date_from = params.get("date_from")
    date_to = params.get("date_to")
    if date_from and date_to:
        try:
            qs = qs.filter(invoice__created_at__range=[date_from, date_to])
        except Exception:
            pass

    aggregated = (
        qs.values("product_id", "product__name", "product__sku")
          .annotate(qty_sold=Sum("quantity"), revenue=Sum(F("subtotal")))
    )

    # --- Nuevo: permitir elegir modo ---
    mode = params.get("mode", "top")  # default "top"
    if mode == "bottom":
        aggregated = aggregated.order_by("qty_sold")  # menos vendidos
    else:
        aggregated = aggregated.order_by("-qty_sold")  # más vendidos

    limit = int(params.get("limit") or 20)

    rows = []
    for a in aggregated[:limit]:
        rows.append({
            "product_id": a["product_id"],
            "sku": a.get("product__sku"),
            "name": a.get("product__name"),
            "qty_sold": int(a.get("qty_sold") or 0),
            "revenue": float(a.get("revenue") or 0),
        })
    columns = ["product_id", "sku", "name", "qty_sold", "revenue"]
    return rows, columns


@ReportService.register("reservations")
def reservations_report(definition: Any, params: Dict[str, Any]):
    """
    Reporte de reservas (apartados).
    Parámetros: date_from, date_to, status
    """
    if Reservation is None:
        raise RuntimeError("Reservation no está disponible.")

    qs = Reservation.objects.all()
    date_from = params.get("date_from")
    date_to = params.get("date_to")
    if date_from and date_to:
        try:
            qs = qs.filter(created_at__range=[date_from, date_to])
        except Exception:
            pass

    if params.get("status"):
        qs = qs.filter(status=params["status"])

    rows = []
    for r in qs.order_by("-created_at"):
        rows.append({
            "id": r.pk,
            "client": f"{r.client_first_name or ''} {r.client_last_name or ''}".strip(),
            "phone": r.client_phone,
            "status": r.status,
            "amount_deposited": float(r.amount_deposited or 0),
            "total": float(r.total or 0) if hasattr(r, "total") else None,
            "remaining_due": float(r.remaining_due or 0),
            "movement_created": bool(getattr(r, "movement_created", False)),
            "days_remaining": r.days_remaining() if hasattr(r, "days_remaining") else None,
            "created_at": getattr(r, "created_at", None).isoformat() if getattr(r, "created_at", None) else None,
            "due_date": getattr(r, "due_date", None).isoformat() if getattr(r, "due_date", None) else None,
        })
    columns = ["id", "client", "phone", "status", "amount_deposited", "total", "remaining_due", "movement_created", "days_remaining", "created_at", "due_date"]
    return rows, columns


@ReportService.register("audit")
def audit_report(definition: Any, params: Dict[str, Any]):
    """
    Auditoría: date_from, date_to
    """
    qs = AuditLog.objects.all().order_by("-created_at")
    date_from = params.get("date_from")
    date_to = params.get("date_to")
    if date_from and date_to:
        try:
            qs = qs.filter(created_at__range=[date_from, date_to])
        except Exception:
            logger.warning("No se pudo aplicar rango de fechas %s - %s", date_from, date_to)

    rows = []
    for a in qs:
        rows.append({
            "id": a.pk,
            "user": getattr(a.user, "username", None),
            "action": a.action,
            "model": a.model,
            "object_id": a.object_id,
            "description": a.description,
            "data": a.data,
            "created_at": a.created_at.isoformat() if getattr(a, "created_at", None) else None,
        })
    columns = ["id", "user", "action", "model", "object_id", "description", "data", "created_at"]
    return rows, columns

@ReportService.register("categories")
def categories_report(definition: Any, params: Dict[str, Any]):
    """
    Ventas agrupadas por categoría.
    Parámetros: date_from, date_to.
    """
    if InvoiceItem is None or not hasattr(Product, "absolute_category"):
        raise RuntimeError("Modelos necesarios no disponibles.")

    qs = InvoiceItem.objects.select_related("product__absolute_category", "invoice")
    if params.get("date_from") and params.get("date_to"):
        qs = qs.filter(invoice__created_at__range=[params["date_from"], params["date_to"]])

    aggregated = (
        qs.values("product__absolute_category__id", "product__absolute_category__name")
          .annotate(total_qty=Sum("quantity"), total_revenue=Sum("subtotal"))
          .order_by("-total_revenue")
    )

    rows = [{
        "category_id": a["product__absolute_category__id"],
        "category_name": a["product__absolute_category__name"],
        "total_qty": int(a["total_qty"] or 0),
        "total_revenue": float(a["total_revenue"] or 0),
    } for a in aggregated]

    columns = ["category_id", "category_name", "total_qty", "total_revenue"]
    return rows, columns


@ReportService.register("daily")
def daily_report(definition: Any, params: Dict[str, Any]):
    """
    Reporte Diario: ventas, reservas y movimientos de inventario de un día.
    Parámetro: date (ISO) -> por defecto hoy.
    """
    if Invoice is None or Reservation is None:
        raise RuntimeError("Billing models no disponibles.")

    date_str = params.get("date") or timezone.now().date().isoformat()
    date = timezone.datetime.fromisoformat(date_str).date()

    # Ventas del día
    sales_qs = Invoice.objects.filter(created_at__date=date)
    total_sales = sales_qs.aggregate(total=Sum("total"))["total"] or 0
    sales_count = sales_qs.count()

    # Reservas del día
    res_qs = Reservation.objects.filter(created_at__date=date)
    reservations_count = res_qs.count()

    rows = [{
        "date": date.isoformat(),
        "sales_count": sales_count,
        "total_sales": float(total_sales),
        "reservations_count": reservations_count,
    }]
    columns = ["date", "sales_count", "total_sales", "reservations_count"]
    return rows, columns

@ReportService.register("monthly")
def monthly_report(definition: Any, params: Dict[str, Any]):
    """
    Reporte Mensual: tendencias de ventas, reservas, top productos.
    Parámetro: month (YYYY-MM), por defecto mes actual.
    """
    if Invoice is None:
        raise RuntimeError("Invoice no disponible.")

    month_str = params.get("month") or timezone.now().strftime("%Y-%m")
    year, month = map(int, month_str.split("-"))
    start = timezone.datetime(year, month, 1, tzinfo=timezone.utc)
    end = (start + timezone.timedelta(days=32)).replace(day=1)

    sales_qs = Invoice.objects.filter(created_at__range=[start, end])
    total_sales = sales_qs.aggregate(total=Sum("total"))["total"] or 0
    sales_count = sales_qs.count()

    rows = [{
        "month": month_str,
        "sales_count": sales_count,
        "total_sales": float(total_sales),
    }]
    columns = ["month", "sales_count", "total_sales"]
    return rows, columns
